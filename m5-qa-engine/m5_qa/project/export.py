"""Project export functions (00108-03).

WHAT: Generates Excel compliance matrices and PDF reports for projects.
      Separated from M8 routes for testability and module independence.

WHY: Business logic belongs in M5, not in the API gateway layer.
     This module can be tested independently without HTTP context.
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Optional

logger = logging.getLogger(__name__)


async def export_compliance_excel(project_manager, project_id: str) -> bytes:
    """Generate an Excel compliance matrix for a project.

    Args:
        project_manager: ProjectManager instance.
        project_id: Target project ID.

    Returns:
        Bytes of the .xlsx file.

    Raises:
        ImportError: If openpyxl is not installed.
        ValueError: If project not found.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        )

    project = await project_manager.get_project(project_id)
    if not project:
        raise ValueError(f"Project not found: {project_id}")

    items = await project_manager.list_compliance(project_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"

    # Headers
    ws.append(["Clause", "Status", "Deviation Note", "Verified By", "Project"])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30

    # Data
    for item in items:
        ws.append([
            item.get("clause_ref", ""),
            item.get("status", "unverified"),
            item.get("deviation_note", ""),
            item.get("verified_by", ""),
            project.get("name", ""),
        ])

    # Conditional formatting
    from openpyxl.styles import PatternFill
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == "verified":
                cell.fill = green_fill
            elif cell.value == "needs_review":
                cell.fill = yellow_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
